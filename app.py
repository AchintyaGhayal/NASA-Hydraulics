

# app.py
# Run: streamlit run app.py

import json, joblib, numpy as np, pandas as pd, matplotlib.pyplot as plt
import streamlit as st
from pathlib import Path
from typing import Dict, List
import base64
import requests
from io import BytesIO
import math

# --- plot style ---
try:
    plt.style.use("seaborn-v0_8")
except Exception:
    plt.style.use("ggplot")

st.set_page_config(page_title="EcoHydra — Green Hydraulics Predictor", layout="wide")

# =======================
# Tiny helpers
# =======================
def _load_metrics():
    p = Path("artifacts/metrics.json")
    if p.exists():
        try:
            return json.loads(p.read_text())
        except Exception:
            return None
    return None

def _fmt(x, nd=3):
    try:
        return f"{float(x):.{nd}f}"
    except Exception:
        return "—"

# =======================
# Theme & header
# =======================
THEME_CHOICE = st.selectbox("Theme", ["Auto", "Light", "Dark"], index=0, key="theme_select", label_visibility="collapsed")

def inject_theme(theme: str):
    st.markdown(f"""
    <style>
      :root {{ --accent:#22c55e; --accent2:#16a34a; }}
      .theme-dark {{
         --bg:#0b1220; --text:#e5e7eb; --muted:#9aa0a6; --card:#111827; --border:rgba(255,255,255,.08);
         --shadow:0 6px 24px rgba(0,0,0,.35);
      }}
      .theme-light {{
         --bg:#f7fafc; --text:#0f172a; --muted:#64748b; --card:#ffffff; --border:rgba(2,6,23,.08);
         --shadow:0 6px 24px rgba(2,6,23,.06);
      }}
      [data-testid="stAppViewContainer"] {{
        background:
          radial-gradient(1200px 600px at 10% -10%, rgba(34,197,94,.14), transparent 60%),
          radial-gradient(1000px 500px at 100% 0%, rgba(20,184,166,.10), transparent 50%),
          var(--bg);
        color: var(--text);
      }}
      .kpi {{ border:1px solid var(--border); border-radius:16px; padding:14px 16px;
             background:linear-gradient(180deg, rgba(255,255,255,.04), rgba(255,255,255,.02)); box-shadow:var(--shadow); }}
      .kpi h3 {{ color: var(--muted); margin:0 0 6px 0; font-size:.95rem; }}
      .kpi div {{ font-weight:700; font-size:1.35rem; }}
      .hero {{ display:flex; align-items:center; justify-content:space-between; gap:16px;
               padding:18px 20px; margin:6px 0 10px; border:1px solid var(--border); border-radius:18px;
               background:linear-gradient(135deg, rgba(34,197,94,.12), rgba(20,184,166,.08)); box-shadow:var(--shadow); }}
      .hero h1 {{ margin:0; }}
      .hero p {{ margin:2px 0 0; color:var(--muted); }}
      .badge {{ display:inline-block; padding:6px 10px; border-radius:999px;
               background:rgba(34,197,94,.15); border:1px solid var(--border); margin-left:6px; }}
      .stButton>button, .stDownloadButton>button {{ border-radius:10px; border:1px solid var(--border); box-shadow:var(--shadow); }}
    </style>
    <script>
      document.documentElement.classList.remove('theme-light','theme-dark');
      {'document.documentElement.classList.add("theme-dark");' if theme=='Dark' else ('document.documentElement.classList.add("theme-light");' if theme=='Light' else '')}
    </script>
    """, unsafe_allow_html=True)

inject_theme(THEME_CHOICE)

st.markdown(
    "**Data:** Powered by [NASA POWER](https://power.larc.nasa.gov/) (T2M, WS10M).",
    help="We use NASA POWER near-surface temperature and wind to ground heat-loss assumptions."
)

st.markdown("""
<div class="hero">
  <div>
    <h1>EcoHydra</h1>
    <p>Green Hydraulics Advisor — efficiency & CO₂ optimization</p>
  </div>
  <div>
    <span class="badge">CERN's Archive</span>
    <span class="badge">ML + Sustainability</span>
  </div>
</div>
""", unsafe_allow_html=True)

# =======================
# Constants & paths
# =======================
CSV_FILE = "hydraulics_ops_sustainability_dataset.csv"
ART = Path("artifacts"); ART.mkdir(exist_ok=True)
MODEL_EFF = ART / "model_efficiency.joblib"
MODEL_CO2 = ART / "model_co2.joblib"
FEATS_EFF = ART / "feature_names_efficiency.json"
FEATS_CO2 = ART / "feature_names_co2.json"
METRICS = ART / "metrics.json"

NUM_EFF = ["filtration_rating_micron","oil_change_interval_hours","test_pressure_bar","test_duration_min","leak_rate_ml_min"]
CAT_COMMON = ["region","client_type","product_line","material","control_type","sensor_pack","oil_type"]
NUM_CO2 = ["energy_used_kwh","filtration_rating_micron","test_pressure_bar","test_duration_min","efficiency_pct"]

# =======================
# Data & model loaders
# =======================
@st.cache_data(show_spinner=False)
def load_reference():
    df = pd.read_csv(CSV_FILE)
    for c in CAT_COMMON:
        if c in df.columns:
            df[c] = df[c].astype("category")
    # Minimal schema check
    required = set(NUM_EFF + NUM_CO2 + CAT_COMMON + ["co2_kg"])
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.warning(f"Dataset missing columns: {missing}")
    return df

@st.cache_resource(show_spinner=False)
def load_models():
    if not (MODEL_EFF.exists() and MODEL_CO2.exists()):
        return None, None
    eff = joblib.load(MODEL_EFF)
    co2 = joblib.load(MODEL_CO2)
    return eff, co2

@st.cache_resource(show_spinner=False)
def load_feature_specs():
    specs = {}
    if FEATS_EFF.exists():
        specs["eff"] = json.loads(FEATS_EFF.read_text())
    if FEATS_CO2.exists():
        specs["co2"] = json.loads(FEATS_CO2.read_text())
    return specs

def train_models_inline():
    # safety net for first run if artifacts missing
    from sklearn.compose import ColumnTransformer
    from sklearn.impute import SimpleImputer
    from sklearn.preprocessing import OneHotEncoder
    from sklearn.pipeline import Pipeline
    from sklearn.ensemble import HistGradientBoostingRegressor

    df = load_reference()

    def build_pre(numeric: List[str], categorical: List[str]):
        num_pipe = Pipeline([("imputer", SimpleImputer(strategy="median"))])
        cat_pipe = Pipeline([
            ("imputer", SimpleImputer(strategy="most_frequent")),
            ("ohe", OneHotEncoder(handle_unknown="ignore", sparse_output=False)),
        ])
        return ColumnTransformer([
            ("num", num_pipe, numeric),
            ("cat", cat_pipe, categorical),
        ], remainder="drop")

    pre1 = build_pre(NUM_EFF, CAT_COMMON)
    m1 = Pipeline([
        ("pre", pre1),
        ("hgb", HistGradientBoostingRegressor(random_state=42, learning_rate=0.08, max_depth=6, max_iter=400, l2_regularization=0.02, early_stopping=True)),
    ])
    X1 = df[NUM_EFF + CAT_COMMON].copy(); y1 = df["efficiency_pct"].copy()
    m1.fit(X1, y1)
    joblib.dump(m1, MODEL_EFF)
    FEATS_EFF.write_text(json.dumps({"numeric": NUM_EFF, "categorical": CAT_COMMON, "target": "efficiency_pct"}, indent=2))

    pre2 = build_pre(NUM_CO2, CAT_COMMON)
    m2 = Pipeline([
        ("pre", pre2),
        ("hgb", HistGradientBoostingRegressor(random_state=42, learning_rate=0.07, max_depth=6, max_iter=450, l2_regularization=0.04, early_stopping=True)),
    ])
    X2 = df[NUM_CO2 + CAT_COMMON].copy(); y2 = df["co2_kg"].copy()
    m2.fit(X2, y2)
    joblib.dump(m2, MODEL_CO2)
    FEATS_CO2.write_text(json.dumps({"numeric": NUM_CO2, "categorical": CAT_COMMON, "target": "co2_kg"}, indent=2))

def _predict(pipe, row_dict: Dict[str, object]):
    X = pd.DataFrame([row_dict])
    return float(pipe.predict(X)[0])

@st.cache_data(show_spinner=False)
def cached_metrics():
    if METRICS.exists():
        return json.loads(METRICS.read_text())
    return None

# =======================
# Header
# =======================
left, right = st.columns([0.8, 0.2])
with left:
    st.title("EcoHydra — Green Hydraulics Advisor")
    st.caption("Predict **Efficiency** and **CO₂** from materials, oils, and controls. Compare **Baseline vs Optimized** in seconds.")
with right:
    if hasattr(st, "popover"):
        with st.popover("About"):
            st.markdown("""
                **EcoHydra** estimates hydraulic **Efficiency (%)** and **CO₂ (kg)** using ML models trained on operations data.  
                • Toggle **Baseline vs Upgraded** scenarios  
                • See **energy/CO₂ deltas** and download a summary  
                • Safety: advisory analytics — **not** a controller  
                • Cite your **IEEE TechRxiv** DOI in the footer
            """)
    else:
        with st.expander("About"):
            st.markdown("""
                **EcoHydra** estimates hydraulic **Efficiency (%)** and **CO₂ (kg)** using ML models.  
                • Baseline vs Upgraded scenarios • Energy/CO₂ deltas • Advisory only • Cite your TechRxiv DOI
            """)

# =======================
# Load data/models
# =======================
# Load pre-trained models instead of training inline
import joblib
from pathlib import Path
import streamlit as st

MODEL_EFF = Path(__file__).with_name("hydraulics_eff_model.pkl")
MODEL_CO2 = Path(__file__).with_name("hydraulics_co2_model.pkl")

try:
    model_eff = joblib.load(MODEL_EFF)
    model_co2 = joblib.load(MODEL_CO2)
    st.success("✅ Pretrained models loaded successfully.")
except Exception as e:
    st.error("❌ Models not available. Please ensure hydraulics_eff_model.pkl and hydraulics_co2_model.pkl are in the repo.")
    st.error(str(e))
    st.stop()

# =======================
# Sidebar inputs
# =======================
st.sidebar.header("Configure a Job")

def cat_select(col, key):
    opts = list(df_ref[col].cat.categories)
    idx = 0 if len(opts) else 0
    return st.sidebar.selectbox(col.replace("_"," ").title(), opts, index=idx, key=key)

# Example defaults (used for Reset)
EXAMPLE = {
    "region": df_ref["region"].cat.categories[0],
    "client_type": df_ref["client_type"].cat.categories[0],
    "product_line": df_ref["product_line"].cat.categories[0],
    "material": df_ref["material"].cat.categories[0],
    "control_type": df_ref["control_type"].cat.categories[0],
    "sensor_pack": df_ref["sensor_pack"].cat.categories[0],
    "oil_type": df_ref["oil_type"].cat.categories[0],
    "filtration_rating_micron": int(df_ref["filtration_rating_micron"].median()),
    "oil_change_interval_hours": int(df_ref["oil_change_interval_hours"].median()),
    "test_pressure_bar": int(df_ref["test_pressure_bar"].median()),
    "test_duration_min": int(df_ref["test_duration_min"].median()),
    "leak_rate_ml_min": float(df_ref["leak_rate_ml_min"].median()),
    "energy_used_kwh": int(df_ref["energy_used_kwh"].median()),
    # Business
    "electricity_price": 0.12,
    "discount_rate": 0.10,
    "capex_per_machine": 8000,
    "machines": 10,
    "adoption_rate": 0.60,
    "cycles_per_year": 200,
    "maintenance_delta": 0.0
}

if st.sidebar.button("🔁 Reset to Example"):
    for k, v in EXAMPLE.items():
        st.session_state[k] = v
    st.rerun()

# categorical
region       = cat_select("region", "region")
client_type  = cat_select("client_type", "client_type")
product_line = cat_select("product_line", "product_line")
material     = cat_select("material", "material")
control_type = cat_select("control_type", "control_type")
sensor_pack  = cat_select("sensor_pack", "sensor_pack")
oil_type     = cat_select("oil_type", "oil_type")

# numeric
filtration_rating_micron  = st.sidebar.number_input("Filtration Rating (micron)", 1, 1000, EXAMPLE["filtration_rating_micron"], key="filtration_rating_micron")
oil_change_interval_hours = st.sidebar.number_input("Oil Change Interval (hours)", 10, 10000, EXAMPLE["oil_change_interval_hours"], key="oil_change_interval_hours")
test_pressure_bar         = st.sidebar.number_input("Test Pressure (bar)", 1, 2000, EXAMPLE["test_pressure_bar"], key="test_pressure_bar")
test_duration_min         = st.sidebar.number_input("Test Duration (min)", 1, 1000, EXAMPLE["test_duration_min"], key="test_duration_min")
leak_rate_ml_min          = st.sidebar.number_input("Leak Rate (ml/min)", 0.0, 500.0, EXAMPLE["leak_rate_ml_min"], key="leak_rate_ml_min")
energy_used_kwh           = st.sidebar.number_input("Planned Energy Used per run (kWh)", 1, 10000, EXAMPLE["energy_used_kwh"], key="energy_used_kwh")

# Business inputs
st.sidebar.markdown("---")
st.sidebar.subheader("Business Inputs")
electricity_price = st.sidebar.number_input("Electricity price ($/kWh)", 0.01, 2.00, EXAMPLE["electricity_price"], step=0.01, key="electricity_price")
discount_rate     = st.sidebar.number_input("Discount rate (0–1)", 0.00, 1.00, EXAMPLE["discount_rate"], step=0.01, key="discount_rate")
capex_per_machine = st.sidebar.number_input("CapEx per upgraded machine ($)", 0, 10_000_000, EXAMPLE["capex_per_machine"], step=100, key="capex_per_machine")
machines          = st.sidebar.number_input("Number of machines", 1, 10000, EXAMPLE["machines"], key="machines")
adoption_rate     = st.sidebar.slider("Adoption rate", 0.0, 1.0, EXAMPLE["adoption_rate"], 0.05, key="adoption_rate")
cycles_per_year   = st.sidebar.number_input("Operating runs per year", 1, 100000, EXAMPLE["cycles_per_year"], key="cycles_per_year")
maintenance_delta = st.sidebar.number_input("Maintenance Δ per machine ($/yr)", -100000.0, 100000.0, EXAMPLE["maintenance_delta"], step=100.0, key="maintenance_delta")

# =======================
# Baseline/Upgraded dicts
# =======================
baseline = dict(
    region=region, client_type=client_type, product_line=product_line, material=material,
    control_type=control_type, sensor_pack=sensor_pack, oil_type=oil_type,
    filtration_rating_micron=filtration_rating_micron, oil_change_interval_hours=oil_change_interval_hours,
    test_pressure_bar=test_pressure_bar, test_duration_min=test_duration_min,
    leak_rate_ml_min=leak_rate_ml_min, energy_used_kwh=energy_used_kwh
)

def smart_upgrade(b: Dict[str, object]) -> Dict[str, object]:
    up = b.copy()
    # Prefer servo / proportional
    cvals = [str(x).lower() for x in df_ref["control_type"].cat.categories]
    if "servo" in cvals:
        up["control_type"] = df_ref["control_type"].cat.categories[cvals.index("servo")]
    elif "proportional" in cvals:
        up["control_type"] = df_ref["control_type"].cat.categories[cvals.index("proportional")]
    # Prefer biodegradable oil
    ovals = [str(x).lower() for x in df_ref["oil_type"].cat.categories]
    if "biodegradable" in ovals:
        up["oil_type"] = df_ref["oil_type"].cat.categories[ovals.index("biodegradable")]
    # Lighter material
    mvals = [str(x).lower() for x in df_ref["material"].cat.categories]
    if str(b.get("material"," ")).lower() in {"steel","carbon steel"}:
        if "composite" in mvals:
            up["material"] = df_ref["material"].cat.categories[mvals.index("composite")]
        elif "aluminum" in mvals or "aluminum alloy" in mvals:
            up["material"] = df_ref["material"].cat.categories[mvals.index("aluminum" if "aluminum" in mvals else "aluminum alloy")]
    # Finer filtration (down to >=3 µm)
    up["filtration_rating_micron"] = max(3, int(round(b.get("filtration_rating_micron", 25) * 0.7)))
    return up

upgraded = smart_upgrade(baseline)

# =======================
# Predictions & deltas
# =======================
eff_base = _predict(model_eff, {
    'filtration_rating_micron': baseline['filtration_rating_micron'],
    'oil_change_interval_hours': baseline['oil_change_interval_hours'],
    'test_pressure_bar': baseline['test_pressure_bar'],
    'test_duration_min': baseline['test_duration_min'],
    'leak_rate_ml_min': baseline['leak_rate_ml_min'],
    'region': baseline['region'], 'client_type': baseline['client_type'], 'product_line': baseline['product_line'],
    'material': baseline['material'], 'control_type': baseline['control_type'], 'sensor_pack': baseline['sensor_pack'], 'oil_type': baseline['oil_type'],
})

eff_up = _predict(model_eff, {
    'filtration_rating_micron': upgraded['filtration_rating_micron'],
    'oil_change_interval_hours': upgraded.get('oil_change_interval_hours', baseline['oil_change_interval_hours']),
    'test_pressure_bar': upgraded.get('test_pressure_bar', baseline['test_pressure_bar']),
    'test_duration_min': upgraded.get('test_duration_min', baseline['test_duration_min']),
    'leak_rate_ml_min': upgraded.get('leak_rate_ml_min', baseline['leak_rate_ml_min']),
    'region': upgraded['region'], 'client_type': upgraded['client_type'], 'product_line': upgraded['product_line'],
    'material': upgraded['material'], 'control_type': upgraded['control_type'], 'sensor_pack': upgraded['sensor_pack'], 'oil_type': upgraded['oil_type'],
})

# CO2: constant input vs same useful output
co2_base_const = _predict(model_co2, {
    'energy_used_kwh': baseline['energy_used_kwh'],
    'filtration_rating_micron': baseline['filtration_rating_micron'],
    'test_pressure_bar': baseline['test_pressure_bar'],
    'test_duration_min': baseline['test_duration_min'],
    'region': baseline['region'], 'client_type': baseline['client_type'], 'product_line': baseline['product_line'],
    'material': baseline['material'], 'control_type': baseline['control_type'], 'sensor_pack': baseline['sensor_pack'], 'oil_type': baseline['oil_type'],
    'efficiency_pct': eff_base,
})

co2_up_const = _predict(model_co2, {
    'energy_used_kwh': baseline['energy_used_kwh'],
    'filtration_rating_micron': upgraded['filtration_rating_micron'],
    'test_pressure_bar': upgraded.get('test_pressure_bar', baseline['test_pressure_bar']),
    'test_duration_min': upgraded.get('test_duration_min', baseline['test_duration_min']),
    'region': upgraded['region'], 'client_type': upgraded['client_type'], 'product_line': upgraded['product_line'],
    'material': upgraded['material'], 'control_type': upgraded['control_type'], 'sensor_pack': upgraded['sensor_pack'], 'oil_type': upgraded['oil_type'],
    'efficiency_pct': eff_up,
})

# Adjust energy so useful output equal: E_up = E_base * (eff_base/eff_up)
adj_energy_same_output = baseline['energy_used_kwh'] * (eff_base / max(eff_up, 1e-6))
co2_up_same_output = _predict(model_co2, {
    'energy_used_kwh': float(adj_energy_same_output),
    'filtration_rating_micron': upgraded['filtration_rating_micron'],
    'test_pressure_bar': upgraded.get('test_pressure_bar', baseline['test_pressure_bar']),
    'test_duration_min': upgraded.get('test_duration_min', baseline['test_duration_min']),
    'region': upgraded['region'], 'client_type': upgraded['client_type'], 'product_line': upgraded['product_line'],
    'material': upgraded['material'], 'control_type': upgraded['control_type'], 'sensor_pack': upgraded['sensor_pack'], 'oil_type': upgraded['oil_type'],
    'efficiency_pct': eff_up,
})

# Waste energy (kWh)
waste_base     = baseline['energy_used_kwh']       * (1 - eff_base/100.0)
waste_up_const = baseline['energy_used_kwh']       * (1 - eff_up/100.0)
waste_up_same  = adj_energy_same_output            * (1 - eff_up/100.0)

# =======================
# KPIs
# =======================
c1, c2, c3, c4 = st.columns(4)
with c1: st.markdown(f'<div class="kpi"><h3>Baseline Efficiency</h3><div>{eff_base:.1f}%</div></div>', unsafe_allow_html=True)
with c2: st.markdown(f'<div class="kpi"><h3>Upgraded Efficiency</h3><div>{eff_up-eff_base:+.1f}% → {eff_up:.1f}%</div></div>', unsafe_allow_html=True)
with c3: st.markdown(f'<div class="kpi"><h3>CO₂ (Const. Input)</h3><div>{co2_base_const:.2f} → {co2_up_const:.2f} kg</div></div>', unsafe_allow_html=True)
with c4: st.markdown(f'<div class="kpi"><h3>CO₂ (Same Output)</h3><div>{co2_base_const:.2f} → {co2_up_same_output:.2f} kg</div></div>', unsafe_allow_html=True)

_m = _load_metrics()
if _m:
    eff_t = _m.get("efficiency", {}).get("test", {})
    co2_t = _m.get("co2", {}).get("test", {})
    st.markdown(
        f"""
        <div style="display:grid; grid-template-columns: repeat(4, minmax(0,1fr)); gap:12px; margin:.25rem 0 1rem;">
          <div class="kpi"><h3>Eff R² (test)</h3><div>{_fmt(eff_t.get('r2'))}</div></div>
          <div class="kpi"><h3>Eff MAE</h3><div>{_fmt(eff_t.get('mae'))}</div></div>
          <div class="kpi"><h3>CO₂ R² (test)</h3><div>{_fmt(co2_t.get('r2'))}</div></div>
          <div class="kpi"><h3>CO₂ MAE</h3><div>{_fmt(co2_t.get('mae'))}</div></div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.info("Train first to see metrics badges (run `python hydraulics.py`).")
st.markdown("<span class='muted'>‘Same output’ normalizes useful work: input energy scales by baseline vs upgraded efficiency.</span>", unsafe_allow_html=True)

# =======================
# Tabs
# =======================
TAB1, TAB2, TAB3, TAB4, TAB5, TAB6 = st.tabs(
    ["Scenario Comparison", "Past Data", "Recommendations", "Admin", "Research Paper", "How to Use"]
)

# ---- NASA climate context (always visible) ----
st.subheader("NASA Climate Context (POWER)")
lat = st.number_input("Latitude", -90.0, 90.0, 29.76)     # Houston example
lon = st.number_input("Longitude", -180.0, 180.0, -95.37)
year = st.number_input("Year", 2000, 2100, 2025)
params = {
    "parameters": "T2M,WS10M",
    "start": f"{year}0101",
    "end":   f"{year}1231",
    "latitude": lat,
    "longitude": lon,
    "format": "JSON",
}
url = "https://power.larc.nasa.gov/api/temporal/daily/point?parameters=T2M,WS10M&start=20240101&end=20241231&latitude=29.76&longitude=-95.38&community=RE&format=JSON"
try:
    r = requests.get(url, params=params, timeout=20)
    r.raise_for_status()
    j = r.json()
    days = j["properties"]["parameter"]["T2M"].keys()
    t2m = j["properties"]["parameter"]["T2M"]
    ws  = j["properties"]["parameter"]["WS10M"]
    df_nasa = pd.DataFrame({
        "date": pd.to_datetime(list(days), format="%Y%m%d"),
        "T2M_C": [t2m[d] for d in days],
        "WS10M_mps": [ws[d] for d in days],
    }).sort_values("date")
    st.line_chart(df_nasa.set_index("date")[["T2M_C","WS10M_mps"]])
    st.caption("NASA POWER daily 2 m air temperature (T2M) and 10 m wind speed (WS10M).")
except Exception as e:
    st.warning(f"NASA POWER fetch failed: {e}")

with TAB1:
    st.subheader("Scenario CO₂ Comparison")
    labels = ["Baseline\n(const input)", "Upgraded\n(const input)", "Upgraded\n(same output)"]
    vals = [co2_base_const, co2_up_const, co2_up_same_output]
    fig, ax = plt.subplots(figsize=(6.6,4))
    ax.bar(labels, vals)
    ax.set_ylabel("Emissions (kg CO₂eq)")
    ax.set_title("Emissions vs Scenario")
    st.pyplot(fig)

    st.markdown("**Details**")
    detail = pd.DataFrame({
        "Scenario": labels,
        "Efficiency (%)": [eff_base, eff_up, eff_up],
        "Energy Used (kWh)": [baseline['energy_used_kwh'], baseline['energy_used_kwh'], adj_energy_same_output],
        "Waste Energy (kWh)": [waste_base, waste_up_const, waste_up_same],
        "CO₂ (kg)": vals,
    })
    st.dataframe(detail, use_container_width=True)
    st.download_button("Download Scenario Summary (CSV)", detail.to_csv(index=False).encode("utf-8"), file_name="scenario_summary.csv")

    # ---- Business View KPIs ----
    st.subheader("Business View")

    # Annualize per machine
    baseline_kwh_per_machine_year = float(baseline['energy_used_kwh']) * cycles_per_year
    upgraded_kwh_per_machine_year = float(adj_energy_same_output) * cycles_per_year
    delta_kwh_per_machine_year = max(baseline_kwh_per_machine_year - upgraded_kwh_per_machine_year, 0.0)

    annual_savings_per_machine = delta_kwh_per_machine_year * electricity_price + maintenance_delta
    machines_upgraded = int(round(machines * adoption_rate))
    fleet_annual_savings = annual_savings_per_machine * machines_upgraded
    total_capex = capex_per_machine * machines_upgraded

    def npv(rate: float, cashflows: list[float]) -> float:
        return sum(cf / ((1 + rate) ** t) for t, cf in enumerate(cashflows))

    def irr(cashflows: list[float], lo=-0.9, hi=1.0, tol=1e-6, maxit=100):
        f_lo = npv(lo + 1e-12, cashflows); f_hi = npv(hi, cashflows)
        if f_lo * f_hi > 0: return None
        for _ in range(maxit):
            mid = (lo + hi) / 2
            f_mid = npv(mid, cashflows)
            if abs(f_mid) < tol: return mid
            if f_lo * f_mid < 0: hi, f_hi = mid, f_mid
            else: lo, f_lo = mid, f_mid
        return mid

    horizon_years = 5
    cashflows = [-total_capex] + [fleet_annual_savings] * horizon_years
    proj_npv = npv(discount_rate, cashflows)
    proj_irr = irr(cashflows)
    payback_months = (12 * total_capex / fleet_annual_savings) if fleet_annual_savings > 0 else float("inf")

    cA, cB, cC, cD = st.columns(4)
    with cA:
        st.markdown(f'<div class="kpi"><h3>Annual $ Savings</h3><div>${fleet_annual_savings:,.0f}</div></div>', unsafe_allow_html=True)
    with cB:
        pb_txt = "—" if not math.isfinite(payback_months) else f"{payback_months:.1f} mo"
        st.markdown(f'<div class="kpi"><h3>Payback</h3><div>{pb_txt}</div></div>', unsafe_allow_html=True)
    with cC:
        st.markdown(f'<div class="kpi"><h3>NPV (5y, {discount_rate:.0%})</h3><div>${proj_npv:,.0f}</div></div>', unsafe_allow_html=True)
    with cD:
        irr_txt = "—" if proj_irr is None else f"{proj_irr:.1%}"
        st.markdown(f'<div class="kpi"><h3>IRR (5y)</h3><div>{irr_txt}</div></div>', unsafe_allow_html=True)

    # ---- Sensitivity (±20%) ----
    st.subheader("Business Sensitivity (±20%)")
    base = fleet_annual_savings
    price_up = delta_kwh_per_machine_year * (electricity_price * 1.2) * machines_upgraded + maintenance_delta * machines_upgraded
    util_up  = (delta_kwh_per_machine_year * 1.2) * electricity_price * machines_upgraded + maintenance_delta * machines_upgraded

    figS, axS = plt.subplots(figsize=(6.5,4))
    axS.bar(["Base", "Price +20%", "Utilization +20%"], [base, price_up, util_up])
    axS.set_ylabel("Annual savings ($)")
    axS.set_title("Sensitivity of Annual Savings")
    st.pyplot(figS)

    # ---- ExecutiveSummary.xlsx export ----
    def build_exec_summary_excel() -> BytesIO | None:
        try:
            # Use openpyxl if available for formulas; fallback to pandas/xlsxwriter if not.
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active; ws.title = "Assumptions"
            ws.append(["Label", "Value"])
            rows = [
                ("Electricity price ($/kWh)", electricity_price),
                ("Discount rate", discount_rate),
                ("Horizon (years)", horizon_years),
                ("CapEx per machine ($)", capex_per_machine),
                ("Machines", machines),
                ("Adoption rate", adoption_rate),
                ("Machines upgraded", machines_upgraded),
                ("Runs per year", cycles_per_year),
                ("Baseline kWh/run", float(baseline['energy_used_kwh'])),
                ("Upgraded kWh/run", float(adj_energy_same_output)),
                ("Δ kWh per machine/year", delta_kwh_per_machine_year),
                ("Maintenance Δ per machine ($/yr)", maintenance_delta),
            ]
            for r in rows: ws.append(list(r))

            ws2 = wb.create_sheet("Savings")
            ws2.append(["Metric", "Value"])
            ws2.append(["Annual savings per machine ($/yr)", delta_kwh_per_machine_year * electricity_price + maintenance_delta])
            ws2.append(["Machines upgraded (count)", machines_upgraded])
            ws2.append(["Fleet annual savings ($/yr)", fleet_annual_savings])
            ws2.append(["Total CapEx ($)", total_capex])
            ws2.append(["Simple payback (months)", None if not math.isfinite(payback_months) else payback_months])

            ws3 = wb.create_sheet("Finance")
            ws3.append(["Year", "Cash Flow ($)"])
            ws3.append([0, -total_capex])
            for y in range(1, horizon_years+1): ws3.append([y, fleet_annual_savings])
            ws3["E1"] = "Discount rate"; ws3["F1"] = discount_rate
            ws3["E2"] = "NPV"; ws3["F2"] = f"=NPV(F1,B3:B{2+horizon_years})+B2"
            ws3["E3"] = "IRR"; ws3["F3"] = f"=IRR(B2:B{2+horizon_years})"

            ws4 = wb.create_sheet("Scenarios")
            ws4.append(["Scenario", "Efficiency (%)", "Energy (kWh/run)"])
            ws4.append(["Baseline", eff_base/100.0, float(baseline['energy_used_kwh'])])
            ws4.append(["Upgraded (same output)", eff_up/100.0, float(adj_energy_same_output)])

            bio = BytesIO(); wb.save(bio); bio.seek(0); return bio
        except Exception:
            # Fallback simple CSV-in-zip if Excel libs missing
            try:
                import zipfile, io
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as z:
                    asum = pd.DataFrame({
                        "Label": ["Electricity price ($/kWh)","Discount rate","Horizon (years)","CapEx per machine ($)","Machines",
                                  "Adoption rate","Machines upgraded","Runs per year","Baseline kWh/run","Upgraded kWh/run",
                                  "Δ kWh per machine/year","Maintenance Δ per machine ($/yr)"],
                        "Value": [electricity_price,discount_rate,horizon_years,capex_per_machine,machines,
                                  adoption_rate,machines_upgraded,cycles_per_year,float(baseline['energy_used_kwh']),
                                  float(adj_energy_same_output),delta_kwh_per_machine_year,maintenance_delta]
                    })
                    z.writestr("Assumptions.csv", asum.to_csv(index=False))
                    sav = pd.DataFrame({
                        "Metric": ["Annual savings per machine ($/yr)","Machines upgraded (count)","Fleet annual savings ($/yr)","Total CapEx ($)","Simple payback (months)"],
                        "Value": [delta_kwh_per_machine_year * electricity_price + maintenance_delta, machines_upgraded, fleet_annual_savings, total_capex,
                                  (None if not math.isfinite(payback_months) else payback_months)]
                    })
                    z.writestr("Savings.csv", sav.to_csv(index=False))
                zip_buf.seek(0)
                return zip_buf
            except Exception:
                return None

    data_blob = build_exec_summary_excel()
    if data_blob:
        st.download_button(
            "Download Executive Summary (Excel/ZIP)",
            data=data_blob,
            file_name="ExecutiveSummary.xlsx" if isinstance(data_blob, BytesIO) else "ExecutiveSummary.zip",
        )
    else:
        st.info("Install `openpyxl` or `xlsxwriter` to enable Excel export.")

with TAB2:
    st.subheader("Historical Insights from Dataset")
    if any(c not in df_ref.columns for c in ["control_type","heat_loss_kwh","energy_used_kwh","co2_kg"]):
        st.info("Add the reference CSV to unlock insights (histograms & trends).")
    else:
        ref = df_ref.copy()
        def group_ctrl(x):
            c = str(x).lower().strip()
            return "Upgraded" if c in {"servo","proportional", "digital"} else "Baseline"
        ref["group"] = ref["control_type"].apply(group_ctrl)

        colA, colB = st.columns(2)
        with colA:
            st.caption("Waste Energy Distribution (Baseline vs Upgraded)")
            fig2, ax2 = plt.subplots(figsize=(6.2,4))
            ax2.hist(ref[ref["group"]=="Baseline"]["heat_loss_kwh"].dropna(), bins=30, alpha=0.65, label="Baseline")
            ax2.hist(ref[ref["group"]=="Upgraded"]["heat_loss_kwh"].dropna(), bins=30, alpha=0.65, label="Upgraded")
            ax2.set_xlabel("Waste Energy per Window (kWh)"); ax2.set_ylabel("Count"); ax2.legend(); ax2.set_title("Waste Energy")
            st.pyplot(fig2)
        with colB:
            st.caption("Emissions vs Energy Used (trend)")
            fig3, ax3 = plt.subplots(figsize=(6.2,4))
            b = ref[ref["group"]=="Baseline"]; u = ref[ref["group"]=="Upgraded"]
            ax3.scatter(b["energy_used_kwh"], b["co2_kg"], alpha=0.65, label="Baseline", s=16)
            ax3.scatter(u["energy_used_kwh"], u["co2_kg"], alpha=0.65, label="Upgraded", s=16)
            x, y = ref["energy_used_kwh"].to_numpy(), ref["co2_kg"].to_numpy()
            if len(x) >= 2:
                m, c = np.polyfit(x, y, 1); xs = np.linspace(float(np.nanmin(x)), float(np.nanmax(x)), 200)
                ax3.plot(xs, m*xs + c, linestyle="--", linewidth=2, label="Trend")
            ax3.set_xlabel("Energy Used (kWh)"); ax3.set_ylabel("Emissions (kg CO₂eq)")
            ax3.set_title("Emissions vs Energy Used"); ax3.legend()
            st.pyplot(fig3)

with TAB3:
    st.subheader("Targeted Recommendations (One change at a time)")
    suggestions = []

    def try_change(ch_name: str, change_fn):
        d = baseline.copy()
        d = change_fn(d)
        e = _predict(model_eff, {
            'filtration_rating_micron': d['filtration_rating_micron'], 'oil_change_interval_hours': d['oil_change_interval_hours'],
            'test_pressure_bar': d['test_pressure_bar'], 'test_duration_min': d['test_duration_min'], 'leak_rate_ml_min': d['leak_rate_ml_min'],
            'region': d['region'], 'client_type': d['client_type'], 'product_line': d['product_line'], 'material': d['material'],
            'control_type': d['control_type'], 'sensor_pack': d['sensor_pack'], 'oil_type': d['oil_type'],
        })
        c = _predict(model_co2, {
            'energy_used_kwh': baseline['energy_used_kwh'], 'filtration_rating_micron': d['filtration_rating_micron'], 'test_pressure_bar': d['test_pressure_bar'],
            'test_duration_min': d['test_duration_min'], 'region': d['region'], 'client_type': d['client_type'], 'product_line': d['product_line'],
            'material': d['material'], 'control_type': d['control_type'], 'sensor_pack': d['sensor_pack'], 'oil_type': d['oil_type'], 'efficiency_pct': e,
        })
        suggestions.append({"Change": ch_name, "ΔEfficiency (pp)": e - eff_base, "ΔCO₂ (kg)": c - co2_base_const})

    try_change("Switch to Servo/Proportional Control", lambda d: {**d, "control_type": upgraded["control_type"]})
    try_change("Use Biodegradable Oil", lambda d: {**d, "oil_type": upgraded["oil_type"]})
    try_change("Finer Filtration (−30%)", lambda d: {**d, "filtration_rating_micron": max(3, int(round(d["filtration_rating_micron"]*0.7)))})

    if str(baseline.get("material"," ")).lower() in {"steel","carbon steel"}:
        try_change("Lighter Material (Composite/Aluminum)", lambda d: {**d, "material": upgraded["material"]})

    imp = pd.DataFrame(suggestions).sort_values(["ΔCO₂ (kg)", "ΔEfficiency (pp)"], ascending=[True, False])
    st.dataframe(imp, use_container_width=True)
    st.markdown("**Recommendation**")
    if not imp.empty:
        top = imp.iloc[0]
        st.write(f"Prioritize: **{top['Change']}** → ΔEfficiency **{top['ΔEfficiency (pp)']:+.2f} pp**, ΔCO₂ **{top['ΔCO₂ (kg)']:+.2f} kg** (const. input)")
    else:
        st.write("Your configuration appears close to optimal given dataset patterns.")

with TAB4:
    st.subheader("Admin & Citation")
    col1, col2, col3 = st.columns([1,1,1])
    with col1:
        if st.button("Re-train models now"):
            with st.spinner("Training models…"):
                train_models_inline()
                st.success("Models retrained. Refresh the page.")
    with col2:
        have_csv = Path(CSV_FILE).exists()
        st.write(f"CSV detected: **{have_csv}** → `{CSV_FILE}`")
        st.write(f"Models present: **{MODEL_EFF.exists() and MODEL_CO2.exists()}**")
        if METRICS.exists():
            st.write("Metrics file:", str(METRICS))
    with col3:
        st.write("**Cite** (replace with your DOI):")
        st.code("Preprint: Data-Driven Sustainability Optimization in Hydraulic Systems — IEEE TechRxiv (DOI: 10.xxxx/techrxiv.xxxxx)")

    st.markdown("---")
    m = cached_metrics()
    if m:
        st.caption("Holdout + 5-fold CV metrics (from training script).")
        st.json(m)

with TAB5:
    st.subheader("Project Paper")
    DOI = "10.5281/zenodo.16879329"  # 🔁 EDIT THIS
    PAPER_URL = "https://zenodo.org/records/16879329"
    st.link_button("Open on CERN's Zenodo", PAPER_URL)

    st.markdown("**How to cite**")
    st.code(
        f"Ghayal, A. (2025). Data-Driven Sustainability Optimization in Hydraulic Systems: "
        f"Efficiency, Emissions, and Material Innovations. IEEE TechRxiv. https://doi.org/{DOI}"
    )

    st.divider()
    st.markdown("**Inline PDF (optional)**")
    candidates = [Path("docs/preprint.pdf"), Path("docs/TechRxiv_Preprint.pdf"), Path("paper.pdf")]
    for p in candidates:
        if p.exists():
            data = p.read_bytes()
            st.download_button("Download PDF", data=data, file_name=p.name, mime="application/pdf")
            b64 = base64.b64encode(data).decode()
            st.markdown(
                f'<iframe src="data:application/pdf;base64,{b64}" width="100%" height="900" type="application/pdf"></iframe>',
                unsafe_allow_html=True,
            )
            break
    else:
        st.info("Place your paper at **docs/preprint.pdf** (or `paper.pdf`) to embed it here.")

with TAB6:
    st.subheader("How to Use")
    st.markdown("""
    1. Set **location** (for NASA POWER context) and choose your **baseline** parameters in the sidebar.  
    2. The app proposes an **upgraded** configuration (servo/proportional, finer filtration, biodegradable oil).  
    3. Compare **CO₂** across scenarios, then review **Business View** for **$ saved, Payback, NPV, IRR**.  
    4. Download the **Executive Summary** and share with stakeholders.  
    5. See **Recommendations** for the single highest-impact change.
    """)

# =======================
# Footers
# =======================
st.markdown("""
---
<span class='subtle'>Disclaimer: This tool provides decision support and estimates. It is **not** a safety interlock. Validate changes locally and ensure compliance with manufacturer guidance and ISO standards.</span>
""", unsafe_allow_html=True)
st.markdown(
    "<hr/><span class='caption'>Limitations: Estimates only; assumptions are user-editable. No PII is collected. NASA POWER data used under its terms.</span>",
    unsafe_allow_html=True
)
